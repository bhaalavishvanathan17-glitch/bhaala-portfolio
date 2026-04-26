from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from models import ContactForm, RegisteredUser
from supabase_client import supabase_admin
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="Bhaala Portfolio API", version="1.0.0")

# ── CORS ── allow all origins so phones/tablets on same WiFi work ──
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Excel file paths ──────────────────────────
USERS_FILE    = Path(__file__).parent / "registered_users.xlsx"
MESSAGES_FILE = Path(__file__).parent / "messages.xlsx"

DARK_RED = PatternFill("solid", fgColor="6C0B0B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=12)
CENTER   = Alignment(horizontal="center")

# ── Email config (loaded from .env) ──────────
GMAIL_SENDER   = os.getenv("GMAIL_SENDER")          # your Gmail address
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASSWORD")    # Gmail App Password (no spaces)
NOTIFY_EMAIL   = "bhaalavishvanathan17@gmail.com"   # where to send notifications


def send_email(subject: str, html_body: str) -> bool:
    """Send a styled HTML email via Gmail SMTP. Returns True on success."""
    if not GMAIL_SENDER or not GMAIL_APP_PASS:
        print("⚠️  Email not configured — set GMAIL_SENDER and GMAIL_APP_PASSWORD in .env")
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"Bhaala Portfolio <{GMAIL_SENDER}>"
        msg["To"]      = NOTIFY_EMAIL

        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_SENDER, GMAIL_APP_PASS)
            server.sendmail(GMAIL_SENDER, NOTIFY_EMAIL, msg.as_string())

        print(f"✅ Email sent → {NOTIFY_EMAIL}")
        return True
    except Exception as e:
        print(f"⚠️  Email send failed: {e}")
        return False


def _registration_email(user: RegisteredUser, timestamp: str) -> str:
    return f"""
    <html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:24px;">
      <div style="max-width:520px;margin:auto;background:#fff;border-radius:10px;
                  box-shadow:0 2px 8px rgba(0,0,0,.12);overflow:hidden;">
        <div style="background:#6C0B0B;padding:20px 28px;">
          <h2 style="color:#fff;margin:0;">🎉 New Registration — Bhaala Portfolio</h2>
        </div>
        <div style="padding:24px 28px;">
          <table style="width:100%;border-collapse:collapse;font-size:15px;">
            <tr><td style="padding:8px 0;color:#555;width:120px;"><b>Full Name</b></td>
                <td style="padding:8px 0;">{user.name}</td></tr>
            <tr style="background:#fafafa;"><td style="padding:8px 6px;color:#555;"><b>Email</b></td>
                <td style="padding:8px 6px;">{user.email}</td></tr>
            <tr><td style="padding:8px 0;color:#555;"><b>Phone</b></td>
                <td style="padding:8px 0;">{user.phone or "—"}</td></tr>
            <tr style="background:#fafafa;"><td style="padding:8px 6px;color:#555;"><b>Registered At</b></td>
                <td style="padding:8px 6px;">{timestamp}</td></tr>
          </table>
        </div>
        <div style="padding:12px 28px;background:#f9f9f9;font-size:12px;color:#aaa;">
          Sent automatically by your Portfolio API
        </div>
      </div>
    </body></html>
    """


def _contact_email(form: ContactForm, timestamp: str) -> str:
    return f"""
    <html><body style="font-family:Arial,sans-serif;background:#f4f4f4;padding:24px;">
      <div style="max-width:520px;margin:auto;background:#fff;border-radius:10px;
                  box-shadow:0 2px 8px rgba(0,0,0,.12);overflow:hidden;">
        <div style="background:#6C0B0B;padding:20px 28px;">
          <h2 style="color:#fff;margin:0;">📬 New Contact Message — Bhaala Portfolio</h2>
        </div>
        <div style="padding:24px 28px;">
          <table style="width:100%;border-collapse:collapse;font-size:15px;">
            <tr><td style="padding:8px 0;color:#555;width:120px;"><b>From</b></td>
                <td style="padding:8px 0;">{form.name}</td></tr>
            <tr style="background:#fafafa;"><td style="padding:8px 6px;color:#555;"><b>Email</b></td>
                <td style="padding:8px 6px;">{form.email}</td></tr>
            <tr><td style="padding:8px 0;color:#555;vertical-align:top;"><b>Message</b></td>
                <td style="padding:8px 0;">{form.message}</td></tr>
            <tr style="background:#fafafa;"><td style="padding:8px 6px;color:#555;"><b>Received At</b></td>
                <td style="padding:8px 6px;">{timestamp}</td></tr>
          </table>
        </div>
        <div style="padding:12px 28px;background:#f9f9f9;font-size:12px;color:#aaa;">
          Sent automatically by your Portfolio API
        </div>
      </div>
    </body></html>
    """


# ── Excel helpers ─────────────────────────────

def _make_styled_sheet(wb, title: str, headers: list, col_widths: list):
    ws = wb.active
    ws.title = title
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = DARK_RED
        cell.alignment = CENTER
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return ws


def get_users_wb():
    """Return (workbook, worksheet) for registered_users.xlsx."""
    if USERS_FILE.exists():
        wb = openpyxl.load_workbook(USERS_FILE)
        return wb, wb.active
    wb = openpyxl.Workbook()
    _make_styled_sheet(wb, "Registered Users",
                       ["Full Name", "Email", "Phone", "Registered At"],
                       [26, 34, 18, 26])
    wb.save(USERS_FILE)
    return wb, wb.active


def get_messages_wb():
    """Return (workbook, worksheet) for messages.xlsx."""
    if MESSAGES_FILE.exists():
        wb = openpyxl.load_workbook(MESSAGES_FILE)
        return wb, wb.active
    wb = openpyxl.Workbook()
    _make_styled_sheet(wb, "Contact Messages",
                       ["Full Name", "Email", "Message", "Received At"],
                       [24, 34, 60, 26])
    wb.save(MESSAGES_FILE)
    return wb, wb.active


# ── Routes ────────────────────────────────────

@app.get("/")
def root():
    return {
        "app": "Bhaala Portfolio API",
        "status": "running ✅",
        "version": "1.0.0",
        "endpoints": {
            "health":        "GET  /api/health",
            "register_user": "POST /api/register-user",
            "contact":       "POST /api/contact",
        },
        "frontend": "http://localhost:5173",
    }


@app.get("/api/health")
def health():
    return {"status": "ok", "message": "Bhaala Portfolio API is running 🚀"}


@app.post("/api/register-user")
async def register_user(user: RegisteredUser):
    """Save a new registered user's details to Excel AND email the owner."""
    now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    # 1. Save to Excel
    try:
        wb, ws = get_users_wb()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and str(row[1]).lower() == user.email.lower():
                return {"success": True, "message": "User already in records."}
        ws.append([user.name, user.email, user.phone, now])
        wb.save(USERS_FILE)
        print(f"✅ User saved to Excel: {user.name} ({user.email})")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel save failed: {e}")

    # 2. Send email notification
    send_email(
        subject=f"🎉 New Registration: {user.name}",
        html_body=_registration_email(user, now),
    )

    return {"success": True, "message": f"Welcome, {user.name}! Registration complete."}


@app.post("/api/contact")
async def contact(form: ContactForm):
    """Save contact message to Excel + Supabase AND email the owner."""
    now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    saved_excel    = False
    saved_supabase = False

    # 1. Save to Excel
    try:
        wb, ws = get_messages_wb()
        ws.append([form.name, form.email, form.message, now])
        wb.save(MESSAGES_FILE)
        saved_excel = True
        print(f"✅ Message saved to Excel: {form.name} ({form.email})")
    except Exception as e:
        print(f"⚠️  Excel save failed: {e}")

    # 2. Save to Supabase (if configured)
    if supabase_admin:
        try:
            result = supabase_admin.table("contacts").insert({
                "name":    form.name,
                "email":   form.email,
                "message": form.message,
            }).execute()
            saved_supabase = bool(result.data)
        except Exception as e:
            print(f"⚠️  Supabase save failed: {e}")

    # 3. Send email notification
    send_email(
        subject=f"📬 New Message from {form.name}",
        html_body=_contact_email(form, now),
    )

    if saved_excel or saved_supabase:
        return {"success": True, "message": "Message received! I'll get back to you soon. 🙏"}

    raise HTTPException(status_code=500, detail="Failed to save message.")
